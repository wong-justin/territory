#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
from time import sleep
import requests
from scrapy.selector import Selector
from flask import Flask, render_template, jsonify, request, send_file
from openpyxl import load_workbook, Workbook
from openpyxl.styles.fills import PatternFill

app = Flask(__name__, static_folder='public', template_folder='views')

app.secret = os.environ.get('SECRET')

@app.route('/')
def homepage():
  return render_template('index.html')

@app.route('/_download')
def download_file():
  
  print('about to  download')
  path = "assets/results.xlsx" 

  return send_file(path, as_attachment=True)

@app.route('/query', methods=['GET', 'POST'])
def recieve_req(): 
   
  data = request.json
  
  request_key(data['requestKey'])
  
  if request_key() == 'sample_key_random_chars_Sgd7rG8Y':
    return jsonify(sample_data())
  
  auth_headers(format_headers(data['headers']))
  
  people = solve_territory(data['coords'])
  if people is False:
    print('there was an error in solving the territory')
    return jsonify(False);
  
  save_terr_results(people)
  return jsonify(people)

def sample_data():
  sample = [
    ('this is', 'some sample', 'data'),
    ('123 Some St', 'John Doe', ''),
    ('123 Some St', 'Jean Doe', ''),
    ('123 Some St', 'Jenny Doe', ''),
    ('1914 Kingdom Wy', 'Some Name', '(012) 345-6789'),
    ('111 Reallyreallyreallyreallylong Rd', 'Another Name', ''),
    ('42 Wallaby Wy', 'P Sherman', ''),
    ('42 Wallaby Wy', 'Nemo Nemo', '')
  ]
  return sample * 10

def request_key(*key):
  '''Get or set request key'''
  request_key.val = key[0] if key else request_key.val
  return request_key.val

def auth_headers(*headers):
  '''Get or set [authenticated] headers for request to reference site'''
  auth_headers.val = headers[0] if headers else auth_headers.val  
  return auth_headers.val

def solve_territory(coords):
  
  latslngs = format_coords(coords)
  num_results = make_shape_query(latslngs)
  
  if num_results > 1000:
    print('probably too many results ... stopping query now')
    return False
  
  RESULTS_PER_PAGE = 25
  total_people = []
  page = 0
  num_left = num_results
  while num_left > 0:
      html = get_result_page(page)
      people = parse_html_result(html)
      total_people.extend(people)
      page += 1
      num_left -= RESULTS_PER_PAGE
      
      print('getting next page')
      sleep(1)
      
  if len(total_people) == num_results:
      return format_results(total_people)
  else:
      print('Stopped because some results missing; probably stopped by bot detection')
      return False
    
def make_shape_query(latslngs):
  
  headers = auth_headers()
  
  url = 'http://www.referenceusa.com/UsWhitePages/MapSearch/AddShape'

  payload = {
    'requestKey': request_key(),
    'latitudes': latslngs[0],
    'longitudes': latslngs[1],
    'name': 'Shape.1',
    'exclude': 'false',
    'type': 'shape'
  }
  
  r = requests.post(url, data=payload, headers=headers)
  
  # if r status is not success, return false
  print(r.text)
  return r.json()['data']['leads']

def get_result_page(page_num):
  
  headers = auth_headers()

  url = 'http://www.referenceusa.com/UsWhitePages/Result/Page'

  payload = {
    'requestKey': request_key(),
    'sort': "",
    'direction': "Ascending",
    'pageIndex': page_num,
    'optionalColumn': ""
  }

  r = requests.post(url, data=payload, headers=headers)
  sleep(1.5)
  return r.text

def parse_html_result(html_text):
    
  selector = Selector(text=html_text)
    
  body = selector.xpath('//tbody[@id="searchResultsPage"]')
  trs = body.xpath('.//tr')

  people = []
  for tr in trs:
      tds = tr.xpath('.//td')

      person = []
      for j in range(1, 3):
          content = tds[j].xpath('.//a/text()').get()     # names are wrapped in link tags
          person.append(content)
      for j in range(3, 7):
          content = tds[j].xpath('.//text()').get() 
          person.append(content)
      people.append(person)

  return people
  
def format_coords(coords_str):
  
  nums = coords_str.split(',')
  nums = [float(''.join(
    [c for c in s if (c.isdigit() or 
                      c == '-' or 
                      c == '.')])) for s in nums]
  lats = []
  longs = []
  for i, n in enumerate(nums):
      if i % 2 == 0:
          longs.append(n)
      else:
          lats.append(n)
  return (lats, longs)

def format_headers(headers_str):
  i = headers_str.index('Host:')
  s = headers_str[i:].strip()
  lines = s.split('\n');
  
  headers = {}
  for li in lines:
    keyval = li.split(': ')
    headers[keyval[0]] = keyval[1]
  return headers

def format_results(people):
  p_by_street = {}
  for p in people:
      name = p[0] + ' ' + p[1]
      phone = '' if p[5] == 'Not Available' else p[5]
      addr = p[2]

      first_space = addr.index(' ')
      street = addr[first_space+1:]
      if street not in p_by_street:
          p_by_street[street] = []
      p_by_street[street].append((addr, name, phone))

  people = []
  for street, ppl in p_by_street.items():
      ppl.sort(key=lambda p: p[0])
      people.extend(ppl)
  return people

def save_terr_results(people):
    '''
    Given a list of people's info, filter and save to spreadsheet.
    '''
    wb = Workbook()
    # wb = load_workbook(filename='assets/results.xlsx') 
    # wb.create_sheet('test', 0)
    
    ws = wb.active

    ws.append(('ADDRESS', 'NAME', 'PHONE'))

    to_fill = False
    grey_fill = PatternFill(start_color='B0B0B0',
                            end_color='B0B0B0',
                            fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF',
                            end_color='FFFFFF',
                            fill_type='solid')       
    num_addrs = 0
    last_addr = ''
    for p in people:
        name = p[1]
        phone = p[2]
        addr = p[0]
        if addr == last_addr:   # don't repeat duplicate info
            addr = ''
            phone = ''
        else:
            last_addr = addr
            num_addrs += 1 
            
            to_fill = not to_fill
        
        row = (addr, name, phone)
        ws.append(row)
        
        fill = grey_fill if to_fill else white_fill
        this_row = ws[ws.max_row]
        for cell in this_row:            
            cell.fill = fill
            
    for col in ('A', 'B', 'C'):
        cells = ws[col]
                
        fitting_length = 1.5 * max( (len(cell.value)) for cell in cells )
        ws.column_dimensions[col].width = fitting_length
        
    wb.save('assets/results.xlsx')
    print(
        'made spreadsheet with {} people and {} addresses'.format(len(people),
                                                                  num_addrs))
if __name__ == '__main__':
  app.run()